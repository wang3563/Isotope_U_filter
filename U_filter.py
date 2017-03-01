#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Fri Feb 24 16:19:37 2017

@author: julianissen
@author: zongyiwang

Python code to filter a specific column and return mean and 2s error of isotope run. 

"""

import openpyxl
import os
import numpy as np

print("Standard U_filter")
file_name = raw_input("Enter the source file name: ")
column_letter = raw_input("What column are we looking at? ")
column_isotope = raw_input("What isotopes are measured in that column?")
column = column_letter+'{}:'+column_letter+'{}'
    
validEntry = False
    
while not validEntry:
    try:
        wb = openpyxl.load_workbook(file_name)
        validEntry = True
        print "Opened workbook"
    except IOError as e:
        print(os.strerror(e.errno)) 
        file_name = raw_input("Enter the correct file name: ")
        print "Opened workbook"

ws = wb.active

def mean(column):
    outlist = []
    for row in ws.iter_rows(column.format(2, ws.max_row - 8)):
        for cell in row:
            value = cell.value
            if value:
                outlist.append(value)
            else:
                outlist.append(np.nan)
    outarray = np.array(outlist, dtype = np.float) 
    outmean = np.nanmean(a = outarray)
    return outmean


def standdev(column):
    outlist = []
    for row in ws.iter_rows(column.format(2, ws.max_row - 8)):
        for cell in row:
            value = cell.value
            if value:
                outlist.append(value)
            else:
                outlist.append(np.nan)
    outarray = np.array(outlist, dtype = np.float) 
    outstanddev = np.nanstd(a = outarray)
    return outstanddev   

def counts(column):
    total_counts = 0
    for row in ws.iter_rows(column.format(2, ws.max_row -8)):
        for cell in row:
            value = cell.value
            if value:
                total_counts += 1
            else:
                continue
    return total_counts
    
                
mean = mean(column)
standdev = standdev(column)
total_counts = counts(column)

criteria = 44 * (standdev / (total_counts ** 0.5))

outlist = []
outcounts = 0
for row in ws.iter_rows(column.format(2, ws.max_row - 8)):
    for cell in row:
        value = cell.value
        if value:
            if abs(value - mean) > criteria:
                outlist.append(np.nan)
            else:
                outlist.append(value)
                outcounts += 1
        else:
            outlist.append(np.nan)
outarray = np.array(outlist, dtype = np.float)
outmean = np.nanmean(a = outarray)
outstanddev = np.nanstd(a = outarray)
err = 2 * (outstanddev / (outcounts ** 0.5))

print column_isotope, "filtered counts:", outcounts
print column_isotope, "filtered mean :", outmean
print column_isotope, "filtered 2s error:", err
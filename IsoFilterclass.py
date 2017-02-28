#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Tue Feb 28 12:42:10 2017

@author: zongyiwang
"""

import openpyxl
import os
import numpy as np

class IsoFilter:
    def __init__(self, filename, columnletter): # input filename and columnletter as strings
        self.column = str(columnletter)+'{}:'+str(columnletter)+'{}'
        self.filename = str(filename)+'.xlsm'
        self.workbook = openpyxl.load_workbook(self.filename)
        self.ws = self.workbook.active
        self.totalCounts = 0
        self.mean = 0 
        self.filteredMean = 0
        self.err = 0
        self.criteria = 0
        
    
    def getMean(self):
        outlist = []
        for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row - 8)):
        #I'd like the G to be able to be inputed as a function object, but can't figure it out
            for cell in row:
                if cell.value:
                    outlist.append(cell.value)
                else:
                    outlist.append(np.nan)
        outarray = np.array(outlist, dtype = np.float) 
        self.mean = np.nanmean(a = outarray)
        
        


    def getStanddev(self):
        outlist = []
      # column = column_letter+'{}:'+column_letter+'{}'
        for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row - 8)):
            for cell in row:
                value = cell.value
                if value:
                    outlist.append(value)
                else:
                    outlist.append(np.nan)
        outarray = np.array(outlist, dtype = np.float) 
        self.standdev = np.nanstd(a = outarray)
        return self.standdev
        #meancol = np.mean(a = outarray)
        #print meancol   
    
    def getCounts(self):
        total_counts = 0
        #column = column_letter+'{}:'+column_letter+'{}'
        for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row -8)):
            for cell in row:
                value = cell.value
                if value:
                    total_counts += 1

        self.totalCounts = total_counts
        return self.totalCounts
        
    def theFilter(self):
        self.criteria = 44 * (self.standdev / self.totalCounts ** 0.5)
        outlist = []
        outcounts = 0
        for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row - 8)):
            for cell in row:
                value = cell.value
                if value:
                    if abs(value - self.mean) > self.criteria:
                        outlist.append(np.nan)
                    else:
                        outlist.append(value)
                        outcounts += 1
                else:
                    outlist.append(np.nan)
        outarray = np.array(outlist, dtype = np.float)
        self.filteredMean = np.nanmean(a = outarray)
        outstanddev = np.nanstd(a = outarray)
        self.err = 2 * (outstanddev / (outcounts ** 0.5))
        #self.criteria = 44 * (self.standdev / self.totalCounts ** 0.5)
        print 'Filtering done! '
        
        
    def __repr__(self):
        
        return 'Filtered mean: '+str(self.filteredMean)+'\n' + \
        'Filtered 2s error: '+ str(self.err)
        
        
     
    
        
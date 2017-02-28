#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Mon Feb 27 16:00:59 2017

@author: zongyiwang
"""

import openpyxl
import os
import numpy as np

def mean(column):
    outlist = []
    
    for row in ws.iter_rows(column.format(2, ws.max_row - 8)):
        #I'd like the G to be able to be inputed as a function object, but can't figure it out
        for cell in row:
            value = cell.value
            if value:
                outlist.append(value)
            else:
                outlist.append(np.nan)
    outarray = np.array(outlist, dtype = np.float) 
    outmean = np.nanmean(a = outarray)
    return outmean


def standdev(column):
    outlist = []
  # column = column_letter+'{}:'+column_letter+'{}'
    for row in ws.iter_rows(column.format(2, ws.max_row - 8)):
        for cell in row:
            value = cell.value
            if value:
                outlist.append(value)
            else:
                outlist.append(np.nan)
    outarray = np.array(outlist, dtype = np.float) 
    outstanddev = np.nanstd(a = outarray)
    return outstanddev
    #meancol = np.mean(a = outarray)
    #print meancol   

def counts(columnra):
    total_counts = 0
    #column = column_letter+'{}:'+column_letter+'{}'
    for row in ws.iter_rows(column.format(2, ws.max_row -8)):
        for cell in row:
            value = cell.value
            if value:
                total_counts += 1
            else:
                continue
    return total_counts
    

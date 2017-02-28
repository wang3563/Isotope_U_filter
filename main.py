#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Mon Feb 27 16:55:13 2017

@author: zongyiwang
"""
import openpyxl
import os
import numpy as np
def main():
    counter = 0
    while counter <= 3:
        print("Standard U_filter")
        file_name = raw_input("Enter the source file name: ")
        column_letter = raw_input("What column are we looking at? ")
        column = column_letter+'{}:'+column_letter+'{}'
        try:
            wb = openpyxl.load_workbook(file_name+'.xlsm')
            validEntry = True
            print "Opened workbook"
        except IOError as e:
            print(os.strerror(e.errno))
            file_name = raw_input("Enter the correct file name: ")
            wb = openpyxl.load_workbook(file_name+'.xlsm')
            print "Opened workbook"

        ws = wb.active
        mean = mean(column)
        standdev = standdev(column)
        total_counts = counts(column)

        criteria = 44 * (standdev / (total_counts ** 0.5))

        outlist = []
        outcounts = 0
        for row in ws.iter_rows(column.format(2, ws.max_row - 8)):
            for cell in row:
                value = cell.value
                if value:
                    if abs(value - mean) > criteria:
                        outlist.append(np.nan)
                    else:
                        outlist.append(value)
                        outcounts += 1
                else:
                    outlist.append(np.nan)
        outarray = np.array(outlist, dtype = np.float)
        outmean = np.nanmean(a = outarray)
        outstanddev = np.nanstd(a = outarray)
        err = 2 * (outstanddev / (outcounts ** 0.5))

        print "Filtered mean :", outmean
        print "Filtered 2s error: ", err
        counter += 1
def mean(column):
    outlist = []
    
    for row in ws.iter_rows(column.format(2, ws.max_row - 8)):
        #I'd like the G to be able to be inputed as a function object, but can't figure it out
        for cell in row:
            value = cell.value
            if value:
                outlist.append(value)
            else:
                outlist.append(np.nan)
    outarray = np.array(outlist, dtype = np.float) 
    outmean = np.nanmean(a = outarray)
    return outmean


def standdev(column):
    outlist = []
  # column = column_letter+'{}:'+column_letter+'{}'
    for row in ws.iter_rows(column.format(2, ws.max_row - 8)):
        for cell in row:
            value = cell.value
            if value:
                outlist.append(value)
            else:
                outlist.append(np.nan)
    outarray = np.array(outlist, dtype = np.float) 
    outstanddev = np.nanstd(a = outarray)
    return outstanddev
    #meancol = np.mean(a = outarray)
    #print meancol   

def counts(columnra):
    total_counts = 0
    #column = column_letter+'{}:'+column_letter+'{}'
    for row in ws.iter_rows(column.format(2, ws.max_row -8)):
        for cell in row:
            value = cell.value
            if value:
                total_counts += 1
            else:
                continue
    return total_counts
    




